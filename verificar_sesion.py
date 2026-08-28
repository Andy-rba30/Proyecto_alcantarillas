#!/usr/bin/env python3
"""
Verificacion independiente de cierre de sesion.

No lee reportes ni mensajes de commit: comprueba el arbol y el tracker.
Correr desde la raiz del repo, sobre un checkout limpio de origin/main:

    git fetch origin && git checkout origin/main
    python3 verificar_sesion.py

Codigo de salida 0 si todo pasa, 1 si algo falla.

Los tres docs/auditorias/auditoria_*.md son evidencia historica: los conteos
que citan son de cuando se hizo la auditoria y NO deben actualizarse. Por eso
el chequeo 2 los excluye.
"""
import re
import subprocess
import sys
from collections import Counter
from pathlib import Path

FALLOS = []
AVISOS = []

# Evidencia historica: no deben reflejar el estado actual.
DOCS_CONGELADOS = {
    "docs/auditorias/auditoria_matematica.md",
    "docs/auditorias/auditoria_sistema.md",
    "docs/auditorias/auditoria_normativa.md",
}


def ok(msg):
    print(f"  [OK]    {msg}")


def fallo(msg):
    print(f"  [FALLO] {msg}")
    FALLOS.append(msg)


def aviso(msg):
    print(f"  [AVISO] {msg}")
    AVISOS.append(msg)


def sh(cmd):
    return subprocess.run(cmd, shell=True, capture_output=True,
                          text=True).stdout.strip()


# ---------------------------------------------------------------- 1. suite
print("\n1. Suite de tests")
salida = subprocess.run([sys.executable, "-m", "pytest", "-q"],
                        capture_output=True, text=True).stdout
m = re.search(r"(\d+) passed(?:, (\d+) skipped)?", salida)
passed = skipped = 0
if not m:
    fallo("no se pudo leer el resultado de pytest")
else:
    passed, skipped = int(m.group(1)), int(m.group(2) or 0)
    if re.search(r"\d+ (failed|error)", salida):
        fallo(f"la suite NO esta verde: {m.group(0)}")
    else:
        ok(f"{passed} passed, {skipped} skipped")

# --------------------------------------------- 2. docs coinciden con la suite
print("\n2. Los documentos vivos citan el numero real de tests")
desfasados = 0
for doc in Path("docs").rglob("*.md"):
    if doc.as_posix() in DOCS_CONGELADOS:
        continue
    texto = doc.read_text(errors="ignore")
    for n in sorted(set(re.findall(r"(\d{3,4}) passed", texto))):
        if int(n) != passed:
            aviso(f"{doc}: dice '{n} passed', la suite da {passed}")
            desfasados += 1
if not desfasados:
    ok("ningun documento vivo cita un conteo desactualizado")

# ------------------------------------------------------------- 3. git limpio
print("\n3. Estado del remoto")
sh("git fetch origin --prune")
ramas = [r.strip() for r in sh("git branch -r").splitlines()
         if r.strip() and "HEAD" not in r]
sobrantes = [r for r in ramas if r != "origin/main"]
for r in sobrantes:
    pendiente = sh(f"git log --oneline origin/main..{r}")
    if pendiente:
        aviso(f"{r}: tiene {len(pendiente.splitlines())} commit(s) fuera de main")
    else:
        fallo(f"{r}: ya fusionada, deberia estar borrada (CLAUDE.md)")
if not sobrantes:
    ok("solo existe origin/main")

if sh("git status --porcelain"):
    fallo("el working tree tiene cambios sin commitear")
else:
    ok("working tree limpio")

# ---------------------------------------------------------------- 4. tracker
print("\n4. Tracker (matriz_cruzada_auditorias.xlsx)")
try:
    import openpyxl
except ImportError:
    openpyxl = None
    fallo("openpyxl no instalado: pip install openpyxl --break-system-packages")

if openpyxl:
    wb = openpyxl.load_workbook(
        "docs/auditorias/matriz_cruzada_auditorias.xlsx")
    ws = wb["Hallazgos"]
    cab = [c.value for c in ws[1]]
    i = {h: n for n, h in enumerate(cab)}
    filas = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]

    if len(filas) != 234:
        fallo(f"{len(filas)} filas, se esperaban 234")
    else:
        ok("234 filas")

    estados = Counter(r[i["Estado"]] for r in filas)
    print(f"          {dict(estados)}")

    # todo lo cerrado cita un commit que existe en main
    huerfanos = []
    for r in filas:
        if str(r[i["Estado"]]).startswith("Cerrado"):
            ref = str(r[i["Commit / PR"]] or "")
            shas = re.findall(r"\b[0-9a-f]{7,40}\b", ref)
            if not shas:
                huerfanos.append(f"{r[0]}: cerrado sin commit citado")
                continue
            for s in shas:
                estado = sh(
                    f"git merge-base --is-ancestor {s} origin/main; echo $?")
                if estado != "0":
                    huerfanos.append(f"{r[0]}: cita {s}, que no esta en main")
    if huerfanos:
        for h in huerfanos:
            fallo(h)
    else:
        ok("todo hallazgo cerrado cita un commit presente en main")

    # las filas sin cluster no vuelven a "Pendiente"
    sin_cluster = [r for r in filas if r[i["Cluster"]] == "\u2014"]
    malos = [r[0] for r in sin_cluster if r[i["Estado"]] == "Pendiente"]
    if malos:
        fallo(f"filas sin cluster de vuelta en Pendiente: {malos}")
    else:
        ok(f"las {len(sin_cluster)} filas sin cluster siguen fuera de la cola")

# ----------------------------------------------------------------- resumen
print("\n" + "=" * 66)
if FALLOS:
    print(f"FALLA: {len(FALLOS)} problema(s). La sesion no esta cerrada.")
elif AVISOS:
    print(f"PASA con {len(AVISOS)} aviso(s) para revisar a mano.")
else:
    print("PASA: sesion cerrada limpiamente.")
print("=" * 66)
sys.exit(1 if FALLOS else 0)
